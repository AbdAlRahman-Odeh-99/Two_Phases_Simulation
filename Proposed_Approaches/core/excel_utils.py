"""
Shared Excel output styling, used by every algorithm/runner file
(gmm_bandit, gmm_submodular, two_stage, and run_proposed_methods.py).

Previously _style_sheet was defined independently in
gmm_2class_bandit_asymmetric.py AND gmm_2class_submodular_asymmetric.py
(identical bodies, just duplicated), and two_stage_runner.py cross-imported
the bandit copy directly (two_stage_asymmetric.py itself never defined its
own). Consolidated here instead, since it's a generic openpyxl formatting
helper with no method-specific logic -- nothing about it is bandit-,
submodular-, or two_stage-specific.

ADDED: add_run_info_sheet, which writes core.logging_utils' provenance
manifest into the workbook as a third sheet. It lives here rather than in
logging_utils so that logging_utils stays free of an openpyxl import (it is
imported by the algorithm modules, which have no business depending on a
spreadsheet library), and because it is the same kind of thing as
_style_sheet: generic openpyxl formatting with no method-specific logic.
"""

from __future__ import annotations

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def serialize_selected_subsets(selected_subsets):
    """Serialize 1-indexed training subsets into one Excel-safe string."""
    return " | ".join(
        ",".join(str(view) for view in subset)
        for subset in selected_subsets
    )


def _style_sheet(ws):
    HEADER_FILL = PatternFill("solid", fgColor="2F5496")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    ALT_FILL = PatternFill("solid", fgColor="DCE6F1")
    BORDER_SIDE = Side(style="thin", color="B8CCE4")
    CELL_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                          top=BORDER_SIDE, bottom=BORDER_SIDE)

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
        for cell in row:
            cell.fill = fill
            cell.border = CELL_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"

    for col in ws.columns:
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0 for cell in col
        )
        width = max(max_len + 4, 14)
        if col[0].value == "Selected Subsets":
            # A training trace can contain thousands of subsets. Keep the
            # workbook usable instead of assigning a many-thousand-character
            # display width; the full value remains in the cell/formula bar.
            width = min(width, 80)
        ws.column_dimensions[get_column_letter(col[0].column)].width = width
    ws.freeze_panes = "A2"


# ─────────────────────────────────────────────────────────────────────────
# Provenance sheet
# ─────────────────────────────────────────────────────────────────────────
def add_run_info_sheet(wb, info_rows, sheet_name="Run Info"):
    """Write (key, value) provenance pairs as a two-column sheet.

    `info_rows` is what RunContext.info_rows() returns: the flattened
    manifest -- git commit and dirty flag, full argv, resolved arguments,
    library versions, hostname, PBS job id, timestamps, row and failure
    counts. The sidecar {run_id}.manifest.json remains the authoritative
    copy; this sheet exists because these workbooks get opened in Excel and
    emailed around, and a result whose provenance is one tab away is a
    result whose provenance actually gets checked.

    Styled deliberately UNLIKE the data sheets: values are left-aligned and
    not number-formatted, because most of them are strings (a commit hash
    centred and rendered as 0.0000 helps nobody).

    Existing sheets are untouched. Called with info_rows=None (no active
    run context) it is a no-op, so a runner used as a library still writes
    exactly the workbook it always did.
    """
    if not info_rows:
        return None
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    HEADER_FILL = PatternFill("solid", fgColor="2F5496")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    KEY_FONT = Font(bold=True)
    WARN_FONT = Font(bold=True, color="9C0006")

    ws.append(["Field", "Value"])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for key, value in info_rows:
        ws.append([str(key), "" if value is None else str(value)])

    for row in ws.iter_rows(min_row=2):
        k_cell, v_cell = row[0], row[1]
        k_cell.font = KEY_FONT
        k_cell.alignment = Alignment(horizontal="left", vertical="top")
        v_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        # A dirty working tree is the one field here that invalidates
        # reproducibility, so it is the one field that gets shouted.
        if str(k_cell.value).endswith("git_dirty") and str(v_cell.value) == "True":
            v_cell.font = WARN_FONT

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 110
    ws.freeze_panes = "A2"
    return ws


def style_and_save(filename, data_sheets, info_rows=None):
    """Apply _style_sheet to `data_sheets`, append the Run Info sheet, save.

    Collapses the four-line load_workbook / _style_sheet / _style_sheet /
    save incantation that every runner repeats verbatim, so adding the
    provenance sheet took one argument instead of four more copies of it.
    """
    from openpyxl import load_workbook

    wb = load_workbook(filename)
    for name in data_sheets:
        if name in wb.sheetnames:
            _style_sheet(wb[name])
    add_run_info_sheet(wb, info_rows)
    wb.save(filename)
    return filename